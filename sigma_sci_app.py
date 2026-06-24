"""

   INTEGRAÇÃO SIGMA → SCI  v3.7  — Domann Contabilidade
   Auto Posto Murungava Ltda | Empresa SCI nº 29

   v3.6: Paleta Domann exata, Código SCI + Nome SCI editáveis com auto-preenchimento

"""

import re, io, csv, json, os, unicodedata, difflib
from datetime import datetime
from collections import defaultdict, Counter

import streamlit as st

# 
# LOGO DOMANN — SVG inline (aproximação fiel da identidade visual)
# 

DOMANN_LOGO_SVG = (
    '<svg viewBox="0 0 330 88" xmlns="http://www.w3.org/2000/svg" width="330" height="88">'
    # Escudo principal — forma heraldica (largo no topo, ponta na base)
    '<path d="M4 4 L4 52 C4 74 37 84 37 84 C37 84 70 74 70 52 L70 4 Z" fill="#2a2828"/>'
    # Borda exterior dourada
    '<path d="M4 4 L4 52 C4 74 37 84 37 84 C37 84 70 74 70 52 L70 4 Z" fill="none" stroke="url(#goldGrad)" stroke-width="2.5"/>'
    # Borda interior (dupla — fidelidade ao logo)
    '<path d="M9 9 L9 51 C9 69 37 78 37 78 C37 78 65 69 65 51 L65 9 Z" fill="none" stroke="#AC996F" stroke-width="0.9" stroke-opacity="0.5"/>'
    # Fundo interno levemente mais claro para contraste das letras
    '<path d="M11 11 L11 50 C11 66 37 75 37 75 C37 75 63 66 63 50 L63 11 Z" fill="#3a3838"/>'
    # Linhas divisoras do grade (cruzes internas)
    '<line x1="37" y1="11" x2="37" y2="73" stroke="#AC996F" stroke-width="1.1" stroke-opacity="0.65"/>'
    '<line x1="11" y1="41" x2="63" y2="41" stroke="#AC996F" stroke-width="1.1" stroke-opacity="0.65"/>'
    # Letras D O M N nos quatro quadrantes — serif elegante
    '<text x="24" y="32" font-family="Georgia,Times New Roman,serif" font-size="17" font-weight="bold" fill="#FFFFFF" text-anchor="middle">D</text>'
    '<text x="50" y="32" font-family="Georgia,Times New Roman,serif" font-size="17" font-weight="bold" fill="#FFFFFF" text-anchor="middle">O</text>'
    '<text x="24" y="65" font-family="Georgia,Times New Roman,serif" font-size="17" font-weight="bold" fill="#FFFFFF" text-anchor="middle">M</text>'
    '<text x="50" y="65" font-family="Georgia,Times New Roman,serif" font-size="17" font-weight="bold" fill="#FFFFFF" text-anchor="middle">N</text>'
    # Degradê dourado para borda do escudo
    '<defs><linearGradient id="goldGrad" x1="0" y1="0" x2="0" y2="1">'
    '<stop offset="0%" stop-color="#C4A96A"/>'
    '<stop offset="50%" stop-color="#AC996F"/>'
    '<stop offset="100%" stop-color="#8B7B4C"/>'
    '</linearGradient></defs>'
    # Nome DOMANN à direita — Inter bold, branco
    '<text x="86" y="44" font-family="Arial Black,Arial,Helvetica,sans-serif" font-size="30" font-weight="900" fill="#FFFFFF" letter-spacing="1">DOMANN</text>'
    # Subtítulo CONTABILIDADE — rastreado, dourado
    '<text x="88" y="62" font-family="Arial,Helvetica,sans-serif" font-size="9" font-weight="400" fill="#AC996F" letter-spacing="4.5">CONTABILIDADE</text>'
    '</svg>'
)

# 
# CONFIGURAÇÃO DA PÁGINA
# 

st.set_page_config(
    page_title="Sigma SCI | Domann Contabilidade",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="collapsed",
)

# Paleta oficial Domann Contabilidade (manual de marca)
_DOMANN_DARK      = "#333232"   # Cinza escuro principal (fundo geral/header)
_DOMANN_DEEP      = "#1e1a1f"   # Cinza profundo (gradiente)
_DOMANN_GOLD      = "#AC996F"   # Dourado claro/champanhe (botões, realces primários)
_DOMANN_GOLD_L    = "#c4b08a"   # Dourado claro hover
_DOMANN_GOLD_S    = "#8B7B4C"   # Dourado fechado/bronze (subtextos, detalhes)
_DOMANN_CREAM     = "#faf8f4"   # Creme (fundo de conteúdo)
_DOMANN_TEXT      = "#333232"   # Texto principal (usa o cinza da marca)
_DOMANN_MUTED     = "#7a7078"   # Texto muted
_DOMANN_BORDER    = "rgba(172,153,111,0.22)"  # Borda sutil dourada

st.markdown(f"""
<style>
/* ── Google Fonts ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Playfair+Display:ital,wght@0,700;1,600&display=swap');

/* ── Base ── */
html, body, [class*="css"] {{
  font-family: 'Inter', system-ui, sans-serif;
  background-color: {_DOMANN_CREAM};
  color: {_DOMANN_TEXT};
}}
.block-container {{ padding:1.2rem 2.2rem 2.5rem; max-width:1500px; }}
h1,h2,h3 {{
  font-family: 'Playfair Display', Georgia, serif;
  color: {_DOMANN_DARK} !important;
  margin-bottom:.5rem;
}}
h3 {{ font-size:1.15rem !important; }}

/* ── Hero ── */
.hero {{
  background: {_DOMANN_DARK};
  border-radius:4px; padding:1.4rem 2rem; margin-bottom:1.8rem;
  display:flex; align-items:center; gap:1.6rem;
  border-bottom:2px solid {_DOMANN_GOLD};
  box-shadow: 0 4px 20px rgba(0,0,0,.4);
}}
.hero-title {{
  font-family: 'Inter', sans-serif; font-size:1.2rem; font-weight:700;
  color:#fff; letter-spacing:.3px;
}}
.hero-sub {{
  color:rgba(255,255,255,.5); font-size:.78rem;
  margin:.2rem 0 0; letter-spacing:.4px;
}}
.hero-tag {{
  display:inline-block; background:{_DOMANN_GOLD}22;
  border:1px solid {_DOMANN_GOLD}88; border-radius:20px;
  padding:3px 12px; font-size:.67rem; color:{_DOMANN_GOLD};
  margin-top:.65rem; letter-spacing:1px; text-transform:uppercase;
  font-weight:700;
}}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {{
  background:#eeebe4; padding:3px; border-radius:5px; gap:2px;
  border:1px solid {_DOMANN_BORDER};
}}
.stTabs [data-baseweb="tab"] {{
  border-radius:4px; padding:.45rem 1.2rem; font-weight:600;
  font-size:.84rem; color:{_DOMANN_MUTED}; border:none !important;
  transition:background .12s;
}}
.stTabs [aria-selected="true"] {{
  background:{_DOMANN_DARK} !important; color:#fff !important;
}}

/* ── Metric cards ── */
.metric-card {{
  background:#fff; border:1px solid {_DOMANN_BORDER}; border-radius:6px;
  padding:1.1rem 1.2rem; text-align:center;
  box-shadow:0 1px 6px rgba(50,45,51,.07);
  transition:box-shadow .15s;
}}
.metric-card:hover {{ box-shadow:0 4px 14px rgba(50,45,51,.13); }}
.metric-value {{ font-size:1.85rem; font-weight:700; color:{_DOMANN_DARK}; line-height:1.1; }}
.metric-label {{ font-size:.7rem; color:{_DOMANN_MUTED}; margin-top:4px; text-transform:uppercase; letter-spacing:.7px; }}
.metric-card-gold {{
  background:linear-gradient(135deg,{_DOMANN_CREAM} 0%,#f0ead9 100%);
  border:1px solid {_DOMANN_GOLD}55; border-radius:6px;
  padding:1.1rem 1.2rem; text-align:center;
  box-shadow:0 1px 6px rgba(184,150,90,.1);
  transition:box-shadow .15s;
}}
.metric-card-gold:hover {{ box-shadow:0 4px 14px rgba(184,150,90,.2); }}
.metric-value-gold {{ font-size:1.85rem; font-weight:700; color:{_DOMANN_GOLD}; line-height:1.1; }}

/* ── Status badges (para renderização HTML inline) ── */
.sbadge {{
  display:inline-flex; align-items:center; gap:5px; padding:4px 11px;
  border-radius:20px; font-size:.73rem; font-weight:700;
  letter-spacing:.2px; white-space:nowrap; line-height:1;
}}
.sbadge::before {{ content:""; display:inline-block; width:6px; height:6px; border-radius:50%; }}
.s-ok    {{ background:#dcfce7; color:#15803d; }} .s-ok::before    {{ background:#15803d; }}
.s-auto  {{ background:#dbeafe; color:#1d4ed8; }} .s-auto::before  {{ background:#1d4ed8; }}
.s-sug   {{ background:#fef9c3; color:#a16207; }} .s-sug::before   {{ background:#a16207; }}
.s-low   {{ background:#ffedd5; color:#c2410c; }} .s-low::before   {{ background:#c2410c; }}
.s-pend  {{ background:#fee2e2; color:#b91c1c; }} .s-pend::before  {{ background:#b91c1c; }}
.s-excl  {{ background:#f1f5f9; color:#64748b; }} .s-excl::before  {{ background:#64748b; }}

/* ── Info / status boxes ── */
.info-box {{
  background:#eff6ff; border-left:3px solid #3b82f6;
  padding:.75rem 1rem; border-radius:0 6px 6px 0;
  font-size:.84rem; margin:.5rem 0; color:#1e40af;
}}
.warn-box {{
  background:#fffbeb; border-left:3px solid #f59e0b;
  padding:.75rem 1rem; border-radius:0 6px 6px 0;
  font-size:.84rem; margin:.5rem 0; color:#78350f;
}}
.ok-box {{
  background:#f0fdf4; border-left:3px solid #22c55e;
  padding:.75rem 1rem; border-radius:0 6px 6px 0;
  font-size:.84rem; margin:.5rem 0; color:#14532d;
}}
.gold-box {{
  background:{_DOMANN_GOLD}0D; border-left:3px solid {_DOMANN_GOLD};
  padding:.75rem 1rem; border-radius:0 6px 6px 0;
  font-size:.84rem; margin:.5rem 0; color:{_DOMANN_TEXT};
}}
.pend-card {{
  background:#fffef7; border:1px solid #d9c46a; border-radius:6px;
  padding:.9rem 1rem; margin:.45rem 0;
}}
.alt-card {{
  background:#fdfdfc; border:1px solid {_DOMANN_BORDER};
  border-radius:6px; padding:.9rem 1rem; margin:.45rem 0;
}}

/* ── Botão principal (ouro) ── */
.stButton>button {{
  background:{_DOMANN_GOLD}; color:{_DOMANN_DEEP};
  border:none; border-radius:5px; font-weight:600;
  padding:.52rem 1.8rem; letter-spacing:.2px;
  font-family:'Inter', sans-serif; font-size:.875rem;
  transition:all .15s ease; box-shadow:0 1px 4px rgba(184,150,90,.25);
}}
.stButton>button:hover {{
  background:{_DOMANN_GOLD_L} !important; color:{_DOMANN_DEEP} !important;
  box-shadow:0 3px 10px rgba(184,150,90,.35) !important;
  transform:translateY(-1px);
}}
.stButton>button:active {{ transform:translateY(0); }}
.stButton>button[kind="secondary"] {{
  background:#fff; color:{_DOMANN_DARK};
  border:1px solid {_DOMANN_BORDER}; box-shadow:none;
}}
.stButton>button[kind="secondary"]:hover {{
  background:{_DOMANN_CREAM} !important; border-color:{_DOMANN_GOLD} !important;
  color:{_DOMANN_DARK} !important; transform:none;
}}

/* ── Download button ── */
.stDownloadButton>button {{
  background:{_DOMANN_DARK}; color:#fff;
  border:1px solid {_DOMANN_GOLD}44; border-radius:5px;
  font-weight:600; font-size:.875rem;
  transition:all .15s ease;
}}
.stDownloadButton>button:hover {{
  background:{_DOMANN_DEEP} !important;
  border-color:{_DOMANN_GOLD}99 !important;
  box-shadow:0 3px 10px rgba(30,26,31,.3) !important;
  transform:translateY(-1px);
}}

/* ── Divisor ouro ── */
.gold-divider {{
  height:1px;
  background:linear-gradient(90deg,transparent,{_DOMANN_GOLD}55,transparent);
  margin:1.5rem 0; border:none;
}}

/* ── Tabela / dataframe ── */
.dataframe thead th {{
  background:{_DOMANN_DARK} !important; color:#fff !important;
  font-size:.81rem; font-weight:600; padding:.55rem .75rem !important;
}}
.dataframe tbody tr:nth-child(even) {{ background:#f8f6f1; }}
.dataframe tbody td {{ font-size:.82rem; padding:.42rem .65rem !important; }}

/* ── Inputs ── */
.stTextInput>div>div>input,
.stTextArea>div>div>textarea {{
  border:1px solid {_DOMANN_BORDER}; border-radius:5px;
  font-family:'Inter', sans-serif; font-size:.875rem;
  color:{_DOMANN_TEXT}; background:#fff;
  transition:border-color .15s;
}}
.stTextInput>div>div>input:focus,
.stTextArea>div>div>textarea:focus {{
  border-color:{_DOMANN_GOLD}88 !important;
  box-shadow:0 0 0 2px {_DOMANN_GOLD}22 !important;
}}
.stSelectbox>div>div {{
  border:1px solid {_DOMANN_BORDER}; border-radius:5px;
}}

/* ── File uploader ── */
[data-testid="stFileUploadDropzone"] {{
  border:2px dashed {_DOMANN_GOLD}55 !important;
  border-radius:6px !important; background:{_DOMANN_CREAM} !important;
  transition:border-color .15s;
}}
[data-testid="stFileUploadDropzone"]:hover {{
  border-color:{_DOMANN_GOLD}aa !important;
}}

/* ── Sidebar ── */
[data-testid="stSidebar"] {{
  background:{_DOMANN_DARK};
  border-right:1px solid {_DOMANN_GOLD}22;
}}
[data-testid="stSidebar"] * {{
  color:#e8e4df !important; font-family:'Inter',sans-serif;
}}
[data-testid="stSidebar"] h3, [data-testid="stSidebar"] strong {{
  color:#fff !important;
}}

/* ── Expander ── */
[data-testid="stExpander"] {{
  border:1px solid {_DOMANN_BORDER} !important; border-radius:6px !important;
}}
[data-testid="stExpander"] summary {{
  font-weight:600; font-size:.875rem; color:{_DOMANN_DARK};
}}

/* ── Progress (conf %) na tabela ── */
[data-testid="stProgress"] > div {{
  background:{_DOMANN_GOLD} !important;
}}
</style>
""", unsafe_allow_html=True)

# 
# PERSISTÊNCIA JSON
# 

_PERSIST_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "de_para_confirmado.json")


def load_persistent() -> dict:
    # 1. session_state (funciona local e nuvem — dentro da sessão)
    if "persistent_map" in st.session_state:
        return dict(st.session_state["persistent_map"])
    # 2. Arquivo local (funciona no computador / VPS)
    if os.path.exists(_PERSIST_FILE):
        try:
            with open(_PERSIST_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                st.session_state["persistent_map"] = data
                return dict(data)
        except Exception:
            pass
    return {}


def save_persistent(mapping: dict) -> None:
    existing = load_persistent()
    existing.update({k: v for k, v in mapping.items() if v})
    # Sempre atualiza session_state (garante cloud)
    st.session_state["persistent_map"] = existing
    # Tenta gravar em arquivo (local/VPS — falha silenciosa na nuvem)
    try:
        with open(_PERSIST_FILE, "w", encoding="utf-8") as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# 
# NORMALIZAÇÃO (definida cedo — usada por detect_class_sci e smart_match)
# 

def _norm(name: str) -> str:
    nfd = unicodedata.normalize("NFD", name)
    s   = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    s   = s.lower()
    s   = re.sub(r"[^a-z0-9\s/]", " ", s)
    return " ".join(s.split())


_STOPWORDS = {
    "de","do","da","dos","das","e","a","o","os","as","em","para","com",
    "por","um","uma","ao","na","no","nos","nas","que","ou","ate","sob",
    "via","se","sua","seu",
}


def _tokens(name: str) -> set[str]:
    return {t for t in _norm(name).split()
            if t not in _STOPWORDS and len(t) >= 3}


# 
# DE/PARA BASE — SIGMA → SCI (Murungava, Empresa 29)
# 

DE_PARA_BASE: dict[str, str] = {
    # ── Caixas ─────────────────────────────────────────────────────────
    "1.1.1.1.1.01": "494018",   # Caixa PDV
    "1.1.1.1.2.1":  "494020",   # Caixa Adm
    "1.1.1.1.2.2":  "494021",   # Sangrias a Baixar
    "1.1.1.1.2.5":  "494022",   # Cofre
    "1.1.1.1.2.6":  "494023",   # Caixa de Troco
    # ── Bancos ──────────────────────────────────────────────────────────
    "1.1.1.2.09.1":  "47597",   # Sicredi
    "1.1.1.2.09.2":  "86",      # CEF
    "1.1.1.2.09.3":  "4588",    # Itaú
    "1.1.1.2.09.11": "94",      # Bradesco
    "1.1.1.2.09.24": "493924",  # Uniprime
    "1.1.1.2.09.26": "4588",    # Itaú Transição (mesma conta SCI)
    # ── Cheques / Boletos / Aplicações ──────────────────────────────────
    "1.1.1.2.13.3": "494024",   # Cheques Devolvidos nos Bancos
    "1.1.1.3.3":    "494026",   # Boletos a compensar
    "1.1.1.3.5":    "494027",   # Capital integralizado Sicredi
    "1.1.1.4.5":    "49085",    # Aplicação Automática Sicredi
    "1.1.1.4.7":    "494098",   # Investimento Sicredi RDC
    # ── Cartões Débito ──────────────────────────────────────────────────
    "1.1.2.01.01.01":"494033",  "1.1.2.01.01.02":"494034",
    "1.1.2.01.01.03":"494035",  "1.1.2.01.01.04":"494036",
    # ── Cartões Crédito ─────────────────────────────────────────────────
    "1.1.2.01.02.01":"494028",  "1.1.2.01.02.04":"100017",
    "1.1.2.01.02.05":"494029",  "1.1.2.01.02.09":"494030",
    "1.1.2.01.02.18":"494031",
    # ── Frotas / Convênios ───────────────────────────────────────────────
    "1.1.2.01.03.09":"494037",  "1.1.2.01.03.12":"494038",
    "1.1.2.01.03.19":"494039",  "1.1.2.01.03.29":"494040",
    "1.1.2.01.03.30":"494041",  "1.1.2.01.03.31":"494042",
    "1.1.2.01.03.32":"494043",  # TICKET LOG (código SCI específico)
    "1.1.2.01.03.34":"494044",  "1.1.2.01.03.35":"494045",
    "1.1.2.01.03.36":"494046",  "1.1.2.01.03.37":"494047",
    "1.1.2.01.03.38":"494048",
    "1.1.2.01.06":  "494123",   # (-) Crédito Celular → Supera
    # ── Clientes ────────────────────────────────────────────────────────
    "1.1.2.2.1":  "494057",     # Faturas a receber c/ boleto
    "1.1.2.2.2":  "494057",     # Notas a cobrar
    "1.1.2.2.4":  "493909",     # Carta Frete a Receber
    "1.1.2.2.32": "494053",     # Promissória a receber
    "1.1.2.2.37": "494053",     # Cheques a Receber
    "1.1.2.2.37.1":"494053",    # Cheques Pré-datados
    "1.1.2.2.37.6":"494054",    # Cheques à vista
    "1.1.2.2.41": "494057",     # Fretes Faturados
    "1.1.2.2.51": "494057",     # Associação Postos O.Frete
    "1.1.2.2.52": "494057",     # Fatura a Receber com Boleto
    "1.1.2.2.68": "494058",     # Cheques Não Recebidos
    "1.1.2.2.69": "494059",     # Cheques Devolvidos Antigos
    "1.1.2.2.98": "494060",     # Inadimplentes Posto
    "1.1.2.2.99": "494061",     # Inadimplentes c/ Virtus
    # ── Funcionários / Terceiros ─────────────────────────────────────────
    "1.1.2.3.1":  "494063",     # Vale Funcionário
    "1.1.2.6.1":  "494064",     # Adiantamento Rodotank
    # ── Estoque / Imobilizado ────────────────────────────────────────────
    "1.1.5.2":    "434",        # Mercadorias para Revenda
    "1.1.5.11":   "299",        # Adiantamento a fornecedores
    "1.2.3.1.17": "494065",     # Benfeitorias em Imóveis
    "1.2.8.3":    "494131",     # (-) Posto Montanha
    # ── Empréstimos ─────────────────────────────────────────────────────
    "2.1.4.12":   "494066",     # Empréstimo Sicredi — Murungava Fev/25
    "2.1.4.13":   "494067",     # Empréstimo Itaú — Murungava Ago/25
    "2.1.4.21":   "494068",     # Loja Murungava
    # ── Passivo — Obrigações ────────────────────────────────────────────
    "2.1.6.01":   "493902",     # Adiantamentos de clientes → Fornecedores Diversos
    "2.1.6.03":   "493902",     # Contas a pagar → Fornecedores Diversos
    "2.1.6.15":   "494073",     # Bonificações a clientes a pagar
    "2.1.6.19":   "494074",     # Ordem de depósito troco
    "2.1.6.27":   "494023",     # Cheques troco PDV
    "2.1.6.28":   "494023",     # Cheques a pagar Adm
    "2.1.6.29":   "494077",     # Aquisição Sobrados
    "2.1.6.30":   "494073",     # Depósito de Cartão
    "2.1.6.31":   "494073",     # Depósitos Clientes
    "2.1.6.32":   "494073",     # Depósitos não identificados
    "2.2.12.2":   "494057",     # (-) Transf. entre unidades a Receber
    "2.2.13.1":   "493902",     # Transf. entre unidades a Pagar
    "2.2.13.2":   "493902",     # MARELUZ R PEREIRA
    # ── Receita (conta transitória obrigatória) ──────────────────────────
    "3.01.1.1":   "159",        # Vendas PDV → Clientes Diversos
    # ── Faltas e Sobras ─────────────────────────────────────────────────
    "3.06.3.7.1": "494114",     # Sobras em caixas
    "3.06.3.7.2": "494113",     # Faltas em caixas
    "3.06.3.7.4": "494113",     # Falta de produtos
    # ── Despesas financeiras ─────────────────────────────────────────────
    "3.07.1.1.3": "5541",       # Tarifas bancárias
    "3.07.1.1.6": "48836",      # Taxas admin de cartões
    "3.07.1.1.8": "48836",      # Taxas antecipação de cartões
    "3.07.1.1.12":"48836",      # Aluguel Máquinas de Cartão
    "3.07.1.1.13":"5541",       # Tarifas Cesta Serviços Bancários
    "3.07.1.1.14":"5541",       # Tarifas com Cobrança/Boleto
    "3.07.1.3.3": "494100",     # Despesa com Bônus Comercial
    # ── Despesas operacionais ────────────────────────────────────────────
    "3.06.1.1.17":"494119",     # Despesas Refeição Funcionário
    "3.06.1.6.26":"494119",     # Utensílios de Cozinha
    "3.06.2.5.25":"494119",     # Cozinha
    "3.06.2.5.32":"494094",     # RodoTank → Fretes
    "3.06.4.9":   "494094",     # Fretes Combustível
    # ── Excluir (contas de controle puro) ────────────────────────────────
    "4.1": "EXCLUIR",   "4.2": "EXCLUIR",
    "4.4": "EXCLUIR",   "4.5": "EXCLUIR",
    "4.6": "EXCLUIR",   "4.7": "EXCLUIR",
}

EXCLUIR_PREFIXOS = ("4.1","4.2","4.4","4.5","4.6","4.7")
EXCLUIR_EXATOS   = {"EXCLUIR"}
ITAU_CODES       = {"1.1.1.2.09.3","1.1.1.2.09.26"}
BANK_SCI_MAP     = {"1.1.1.2.09.1":"47597","1.1.1.2.09.24":"493924","1.1.1.1.2.5":"494022"}

FUNCIONARIOS_REG: set[str] = {
    "ALMIR SCHUERTS","CRISTIANO NAKONECZYI","DEOMAR COSTA",
    "JEAN CARLOS DE LIMA IRENO","JEFERSON JANSEN","PAULO CEZAR CARLOTTO",
    "ANA CAROLINA KAMINSKI","ANA LETICIA CORREA","ANTONIO BORGES",
    "CARLOS BARBOSA SIEBRE","CARLOS JOCIEL NAIRNE","ANDREY BATISTA LOPES",
    "CLEVERSON PATRIEL NAIRNE","DIRLEI CARLOTTO","ELIANE PACHKO",
    "FABIANO BUDNIAK","GABRIEL VINICIOS LOPES","GISLAINE SOCHA",
    "HADRIAN RAFAEL BATISTA","JANETE PINGAS","LUCAS GABRIEL FELIZ",
    "LUCIA DE OLIVEIRA","MARIA DIATCHUK","MILENY IRENO DE LIMA",
    "MONICA DE CAMARGO ROCHA AMARAL","SHEILA GISELE SOCHA","SOLANGE KAVA",
    "STEFFANY EDUARDA GRESKIW DE OLIVEIRA","VANDERSON SCHUERTS",
    "WITOR LUCIANO MARCAL","DANIELE TUARD","LUAN FELIZ",
    "SOLANGE IRENO DE LIMA","ELIZEU VOGENEY VALES",
}
SOCIOS: set[str] = {
    "MARELUZ RIBEIRO PEREIRA","MARILUZ RIBEIRO PEREIRA",
    "SERGIO AFONSO FELIPPE FILHO",
}

# 
# CLASSIFICAÇÃO PATRIMONIAL
# 

def detect_class_sigma(code: str) -> str:
    c = code.strip()
    if c.startswith("1."): return "ATIVO"
    if c.startswith("2."): return "PASSIVO"
    if re.match(r"^3\.0[1-5]", c): return "RECEITA"
    if re.match(r"^3\.0[6-9]|^3\.[1-9]", c): return "DESPESA"
    if c.startswith("3."): return "RESULTADO"
    return "OUTROS"


_KW_ATIVO   = ["receber","caixa","banco","aplicacao","deposito","estoque",
               "adiantamento","ativo","almoxarifado","frota","credito a recuperar",
               "imposto a recuperar","carta frete a receber"]
_KW_PASSIVO = ["pagar","recolher","passivo","obrigacao","fornecedor","emprestimo",
               "financiamento","retencao","salario a","fgts a","inss a",
               "ferias a","iss a","csll a","irpj a","cofins a","pis a","icms a",
               "troco a pagar"]
_KW_PL      = ["capital","reserva","lucro acumulado","prejuizo acumulado","patrimonio"]
_KW_RECEITA = ["receita","venda","faturamento","honorario recebido","clientes diversos"]
_KW_DESPESA = ["despesa","custo","gasto","taxa","tarifa","comissao","encargo",
               "juro","multa","amortizacao","depreciacao"]


def detect_class_sci(name: str) -> str:
    # Sufixo explícito do parser CSV Domann: 'Nome [Ativo]', 'Nome [Passivo]', etc.
    m = re.search(r'\[(Ativo|Passivo|Receita|Despesa)\]$', name, re.I)
    if m:
        t = m.group(1).upper()
        if t == "ATIVO":   return "ATIVO"
        if t == "PASSIVO": return "PASSIVO"
        if t == "RECEITA": return "RECEITA"
        if t == "DESPESA": return "DESPESA"
    n = _norm(name)
    for kw in _KW_ATIVO:
        if kw in n: return "ATIVO"
    for kw in _KW_PASSIVO:
        if kw in n: return "PASSIVO"
    for kw in _KW_PL:
        if kw in n: return "PL"
    for kw in _KW_RECEITA:
        if kw in n: return "RECEITA"
    for kw in _KW_DESPESA:
        if kw in n: return "DESPESA"
    return "OUTROS"


def classes_compativeis(cls_sigma: str, cls_sci: str) -> bool:
    if cls_sci == "OUTROS" or cls_sigma == "OUTROS": return True
    if cls_sigma == "RESULTADO": return cls_sci in ("RECEITA","DESPESA","RESULTADO","OUTROS")
    return cls_sigma == cls_sci


#  Bloqueio de instituição financeira 
# Impede que o fuzzy-match cruze contas de instituições diferentes.
# Ex: "Aplicação Sicredi" NUNCA pode sugerir "Aplicação Ailos".
_INST_GROUPS: list[list[str]] = [
    ["sicredi"],
    ["ailos"],
    ["xp invest", "xp investimento"],
    ["cef", "caixa econom"],          # CEF / Caixa Econômica
    ["itau", "itaú"],
    ["bradesco"],
    ["banco do brasil", "bb renda", "bb cdb"],
    ["santander"],
    ["sicoob"],
    ["uniprime"],
    ["unicredi"],
    ["sisprime"],
    ["inter"],
    ["c6"],
    ["safra"],
    ["mercado pago"],
    ["repom"],
    ["ticket log"],
    ["ticket "],
    ["goodcard"],
    ["hipercard"],
    ["elo "],
    ["maestro"],
    ["visa"],
    ["mastercard"],
    ["amex", "american express"],
    ["ipiranga"],
    ["cabal"],
]

def _inst_group(name: str) -> str | None:
    """Retorna o grupo de instituição detectado no nome (lower), ou None."""
    n = _norm(name)
    for grp in _INST_GROUPS:
        if any(kw in n for kw in grp):
            return grp[0]   # chave canônica do grupo
    return None


# 
# PARSERS MULTI-FORMATO (Sigma)
# 

_ROW_RE  = re.compile(r"<(?:\w+:)?Row[^>]*>(.*?)</(?:\w+:)?Row>",
                      re.DOTALL | re.IGNORECASE)
_DATA_RE = re.compile(r"<(?:\w+:)?Data[^>]*>([^<]*)</(?:\w+:)?Data>", re.I)
_CELL_RE = re.compile(
    r"<(?:\w+:)?(?:Data|Cell)[^>]*>([^<]*)</(?:\w+:)?(?:Data|Cell)>", re.I)
_SALDO_RE = re.compile(r"Saldo\s+[Ii]nicial", re.I)
_DATE_RE  = re.compile(r"^\d{2}/\d{2}/\d{4}$")
#  FIX: regex para remover prefixo de código do nome 
_CODE_PREFIX_RE = re.compile(r"^\d[\d.]*\.\s*")


def _strip_code_from_name(raw: str) -> str:
    """Remove prefixo de código hierárquico do nome.
    Ex: '1.1.1.2.2. Sangrias a baixar' → 'Sangrias a baixar'
    """
    cleaned = _CODE_PREFIX_RE.sub("", raw).strip()
    return cleaned if cleaned else raw


def _decode_bytes(content: bytes) -> str:
    if content[:2] in (b'\xff\xfe', b'\xfe\xff'):
        return content.decode("utf-16", errors="replace")
    if content[:3] == b'\xef\xbb\xbf':
        return content[3:].decode("utf-8", errors="replace")
    head = content[:300].decode("ascii", errors="replace")
    enc_m = re.search(r'encoding=["\']([^"\']+)["\']', head, re.I)
    if enc_m:
        enc = enc_m.group(1).replace("-","").lower()
        try: return content.decode(enc, errors="replace")
        except Exception: pass
    # Para XML/HTML sem declaração de encoding: tenta UTF-8 estrito primeiro
    # (arquivos SpreadsheetML do FastReport são UTF-8 sem BOM)
    is_xml = b'<?xml' in content[:50] or b'<Workbook' in content[:200]
    if is_xml:
        try: return content.decode("utf-8")
        except UnicodeDecodeError: pass
    for enc in ("windows-1252", "latin-1", "utf-8"):
        try: return content.decode(enc, errors="replace")
        except Exception: pass
    return content.decode("utf-8", errors="replace")


def _detect_format(content: bytes, filename: str) -> str:
    ext = filename.lower().rsplit(".",1)[-1] if "." in filename else ""
    if content[:2] == b"PK": return "xlsx"
    head = content[:1024].lower()
    if (b"<?xml" in head or b"<workbook" in head or b"<ss:workbook" in head
            or b"<row" in head or b"ss:row" in head
            or (ext in ("xls","xml") and b"<" in head)):
        return "xml"
    if b"<html" in head or b"<table" in head or b"<!doctype" in head: return "html"
    if ext == "csv": return "csv"
    if ext == "xls": return "xml"
    return "csv"


def _extract_name_from_cells(cells: list[str]) -> str:
    """Extrai nome da conta de uma linha de cabeçalho, removendo prefixo de código."""
    name_raw = next(
        (c for c in cells
         if not _SALDO_RE.search(c)
         and not re.match(r"^[\d.]+\.?$", c.strip())),
        cells[0]
    )
    return _strip_code_from_name(name_raw)


def _parse_xml(content: bytes) -> dict:
    text = _decode_bytes(content)
    accounts: dict = {}
    current_code = None
    for raw_row in _ROW_RE.findall(text):
        cells = [c.strip() for c in _DATA_RE.findall(raw_row)]
        if not cells: cells = [c.strip() for c in _CELL_RE.findall(raw_row)]
        cells = [c for c in cells if c]
        if not cells: continue
        saldo_idx = next((i for i,c in enumerate(cells) if _SALDO_RE.search(c)), -1)
        if saldo_idx >= 0:
            m = re.match(r"^([\d.]+)", cells[0].strip())
            current_code = m.group(1).rstrip(".") if m else cells[0].strip()
            name = _extract_name_from_cells(cells)
            accounts[current_code] = {"name":name,"saldo_inicial":0.0,"movements":[]}
        elif current_code and cells and _DATE_RE.match(cells[0]):
            try: valor = float(cells[-1].replace(".","").replace(",","."))
            except Exception: valor = 0.0
            accounts[current_code]["movements"].append({
                "date":cells[0],"desc":cells[1] if len(cells)>1 else "",
                "doc":cells[2] if len(cells)>2 else "","valor":valor})
    return accounts


def _parse_html(content: bytes) -> dict:
    try: from bs4 import BeautifulSoup
    except ImportError: return {}
    soup = BeautifulSoup(_decode_bytes(content), "lxml")
    accounts: dict = {}
    current_code = None
    for row in soup.find_all("tr"):
        cells = [td.get_text(strip=True) for td in row.find_all(["td","th"])]
        if not cells: continue
        saldo_idx = next((i for i,c in enumerate(cells) if _SALDO_RE.search(c)), -1)
        if saldo_idx >= 0:
            m = re.match(r"^([\d.]+)", cells[0])
            current_code = m.group(1).rstrip(".") if m else cells[0]
            name = _extract_name_from_cells(cells)
            accounts[current_code] = {"name":name,"saldo_inicial":0.0,"movements":[]}
        elif current_code and cells and _DATE_RE.match(cells[0]):
            try: valor = float(cells[-1].replace(".","").replace(",","."))
            except Exception: valor = 0.0
            accounts[current_code]["movements"].append({
                "date":cells[0],"desc":cells[1] if len(cells)>1 else "",
                "doc":cells[2] if len(cells)>2 else "","valor":valor})
    return accounts


def _parse_xlsx(content: bytes) -> dict:
    try: import openpyxl
    except ImportError: return {}
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    accounts: dict = {}
    current_code = None
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            cells = [c for c in cells if c]
            if not cells: continue
            saldo_idx = next((i for i,c in enumerate(cells) if _SALDO_RE.search(c)), -1)
            if saldo_idx >= 0:
                m = re.match(r"^([\d.]+)", cells[0])
                current_code = m.group(1).rstrip(".") if m else cells[0]
                name = _extract_name_from_cells(cells)
                accounts[current_code] = {"name":name,"saldo_inicial":0.0,"movements":[]}
            elif current_code and cells and _DATE_RE.match(cells[0]):
                try: valor = float(str(cells[-1]).replace(".","").replace(",","."))
                except Exception: valor = 0.0
                accounts[current_code]["movements"].append({
                    "date":cells[0],"desc":cells[1] if len(cells)>1 else "",
                    "doc":cells[2] if len(cells)>2 else "","valor":valor})
    return accounts


def _parse_csv(content: bytes) -> dict:
    text = _decode_bytes(content)
    accounts: dict = {}
    cc = "CSV_IMPORT"
    accounts[cc] = {"name":"Importação CSV","saldo_inicial":0.0,"movements":[]}
    try: dialect = csv.Sniffer().sniff(text[:2048], delimiters=";,\t|")
    except Exception: dialect = csv.excel
    for row in csv.reader(io.StringIO(text), dialect):
        if not row or not _DATE_RE.match(row[0].strip()): continue
        try: valor = float(row[-1].strip().replace(".","").replace(",","."))
        except Exception: valor = 0.0
        accounts[cc]["movements"].append({
            "date":row[0].strip(),"desc":row[1].strip() if len(row)>1 else "",
            "doc":row[2].strip() if len(row)>2 else "","valor":valor})
    return accounts


def parse_sigma_file(file_bytes: bytes, filename: str) -> dict:
    fmt = _detect_format(file_bytes, filename)
    order = {
        "xlsx":[_parse_xlsx,_parse_xml,_parse_html],
        "xml": [_parse_xml,_parse_html,_parse_xlsx,_parse_csv],
        "html":[_parse_html,_parse_xml,_parse_xlsx,_parse_csv],
        "csv": [_parse_csv,_parse_xml,_parse_html],
    }.get(fmt,[_parse_xml,_parse_html,_parse_xlsx,_parse_csv])
    for parser in order:
        try:
            r = parser(file_bytes)
            if r: return r
        except Exception: continue
    return {}


# 
# PARSER PLANO DE CONTAS SCI
# 

def parse_sci_plan(file_bytes: bytes, filename: str) -> dict[str, str]:
    sci: dict[str,str] = {}
    ext = filename.lower().rsplit(".",1)[-1] if "." in filename else ""

    def _try_row(row: list[str]) -> tuple[str,str] | None:
        row = [c.strip() for c in row if c.strip()]
        if len(row) < 2: return None
        for i, cell in enumerate(row):
            if re.match(r"^\d{2,8}$", cell):
                name_parts = [row[j] for j in range(len(row)) if j != i and row[j]]
                return cell, " — ".join(name_parts[:2])
        return None

    def _parse_sci_csv_domann(text: str) -> dict[str, str]:
        """
        Parser específico para o formato de exportação do SCI (Domann).
        Estrutura por conta (blocos de 6 linhas):
          L+0: ';CÓDIGO;CLASSIFICAÇÃO;...'
          L+1: '    Nome da Conta;;;...'
          L+2: '[Apelido?];;;;;;;Tipo(Ativo|Passivo|Receita|Despesa);...'
        Retorna: código → 'Nome [Tipo]' para uso no motor de matching
        """
        result: dict[str,str] = {}
        lines = text.splitlines()
        i = 0
        while i < len(lines):
            parts = lines[i].split(';')
            if len(parts) >= 2:
                code = parts[1].strip()
                if re.match(r'^\d{2,8}$', code) and i+1 < len(lines):
                    name_raw = lines[i+1].split(';')[0].strip()
                    if name_raw and not re.match(r'^\d', name_raw):
                        # Extrai tipo patrimonial da linha +2, posição 7
                        tipo = ''
                        if i+2 < len(lines):
                            t_parts = lines[i+2].split(';')
                            if len(t_parts) > 7:
                                tipo = t_parts[7].strip()
                        # Armazena nome + sufixo de tipo para auxiliar detect_class_sci
                        if tipo in ('Ativo','Passivo','Receita','Despesa'):
                            result[code] = f"{name_raw} [{tipo}]"
                        else:
                            result[code] = name_raw
            i += 1
        return result

    if ext in ("xlsx","xls") and file_bytes[:2] == b"PK":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws_main = wb.active
            # Parser específico para o formato SCI (Domann):
            # Col B=código, Col D=tipo (T=sintética), Col E=nome, Col G=grupo
            for row in ws_main.iter_rows(values_only=True):
                code_raw = row[1] if len(row) > 1 else None
                tipo     = row[3] if len(row) > 3 else None
                name_raw = row[4] if len(row) > 4 else None
                group    = row[6] if len(row) > 6 else None
                if code_raw is None or tipo == 'T': continue
                if not isinstance(code_raw, (int, float)): continue
                code_str = str(int(code_raw))
                if not name_raw: continue
                clean = str(name_raw).strip()
                if group in ('Ativo','Passivo','Receita','Despesa'):
                    sci[code_str] = f"{clean} [{group}]"
                else:
                    sci[code_str] = clean
            # Fallback: se parser específico não achou nada, usa genérico
            if not sci:
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c).strip() if c is not None else "" for c in row]
                        r = _try_row(cells)
                        if r: sci[r[0]] = r[1]
        except Exception: pass
    else:
        # Tenta decodificar — preferência: windows-1252 para CSV do SCI
        for enc in ("windows-1252", "utf-8-sig", "utf-8", "latin-1"):
            try: text = file_bytes.decode(enc); break
            except Exception: pass
        else:
            text = file_bytes.decode("windows-1252", errors="replace")

        # 1º: parser específico do SCI Domann (formato ';código;classif' + nome na próxima linha)
        domann_result = _parse_sci_csv_domann(text)
        if domann_result:
            sci = domann_result
        else:
            # Fallback: parser genérico por delimitador
            best: dict[str,str] = {}
            for delim in (";", ",", "\t", "|", " "):
                tmp: dict[str,str] = {}
                for line in text.splitlines():
                    r = _try_row(line.split(delim))
                    if r: tmp[r[0]] = r[1]
                if len(tmp) > len(best): best = tmp
            sci = best
    return sci


# 
# MOTOR DE MATCHING INTELIGENTE (4 estratégias + lock de classe)
# 

def smart_match(
    sigma_name: str,
    sci_plan: dict[str,str],
    sigma_class: str = "OUTROS",
    top_n: int = 5,
) -> list[tuple[str,str,float,str]]:
    norm_s    = _norm(sigma_name)
    tok_s     = _tokens(sigma_name)
    inst_s    = _inst_group(sigma_name)   # instituição detectada no nome Sigma
    results: list[tuple[str,str,float,str]] = []

    for code, name in sci_plan.items():
        if not classes_compativeis(sigma_class, detect_class_sci(name)): continue
        norm_sci = _norm(name)
        tok_sci  = _tokens(name)
        inst_sci = _inst_group(name)      # instituição detectada no nome SCI

        #  Bloqueio de instituição 
        # Se Sigma menciona instituição X e SCI menciona instituição Y ≠ X
        # → score zero. Jamais sugerir "Ailos" para "Sicredi", etc.
        if inst_s and inst_sci and inst_s != inst_sci:
            continue
        # Bônus: mesma instituição nas duas pontas → multiplica score
        inst_bonus = 1.20 if (inst_s and inst_sci and inst_s == inst_sci) else 1.0

        score, method = 0.0, ""
        if norm_s == norm_sci:
            score, method = 1.00, "Exato"
        elif len(norm_s) >= 4 and norm_s in norm_sci:
            score = 0.88 + len(norm_s)/max(len(norm_sci),1)*0.08
            method = "Contido em SCI"
        elif len(norm_sci) >= 4 and norm_sci in norm_s:
            score = 0.84 + len(norm_sci)/max(len(norm_s),1)*0.08
            method = "SCI contido"
        else:
            tok_score = 0.0
            if tok_s and tok_sci:
                common = tok_s & tok_sci
                if common:
                    prec = len(common)/len(tok_s)
                    rec  = len(common)/len(tok_sci)
                    tok_score = 2*prec*rec/(prec+rec)*0.85
            seq_score = difflib.SequenceMatcher(None,norm_s,norm_sci).ratio()*0.80
            score  = max(tok_score, seq_score)
            method = "Tokens" if tok_score >= seq_score else "Fuzzy"

        score = min(score * inst_bonus, 1.0)
        if score >= 0.25:
            results.append((code, name, round(score,4), method))

    results.sort(key=lambda x:x[2], reverse=True)
    return results[:top_n]


# 
# CONSTRUTOR DO DE/PARA
# 

def build_depara_with_matches(accounts, sci_plan, persistent, extra) -> list[dict]:
    rows = []
    for key, data in sorted(accounts.items()):
        nc    = get_sigma_code(key)
        name  = data.get("name", key)
        movs  = data.get("movements", [])
        total = sum(abs(m.get("valor",0) or 0) for m in movs)
        excluir   = is_excluir(nc)
        cls_sigma = detect_class_sigma(nc)

        sci_code, source = "", ""
        # persistent e extra têm prioridade máxima — inclusive sobre excluir
        if nc in extra and extra[nc]:
            sci_code, source = extra[nc],"sessão"
            excluir = False
        elif nc in persistent and persistent[nc]:
            sci_code, source = persistent[nc],"salvo"
            excluir = persistent[nc] == "EXCLUIR"
        elif excluir:
            sci_code, source = "EXCLUIR","base"
        elif nc in DE_PARA_BASE and DE_PARA_BASE[nc]:
            sci_code, source = DE_PARA_BASE[nc],"base"
        else:
            for k,v in DE_PARA_BASE.items():
                if nc.startswith(k+".") and v:
                    sci_code, source = v,"base-prefix"; break

        sci_name = sci_plan.get(sci_code,"") if (sci_plan and sci_code
                                                  and sci_code != "EXCLUIR") else ""
        candidates = smart_match(name, sci_plan, cls_sigma, top_n=5) if sci_plan else []
        auto_code = auto_name = auto_method = ""
        auto_score = 0.0
        if candidates: auto_code,auto_name,auto_score,auto_method = candidates[0]

        if excluir or sci_code == "EXCLUIR":
            status, conf = "Excluído", 1.0
        elif sci_code and source in ("base","salvo","base-prefix","sessão"):
            status, conf = "Confirmado", 1.0
        elif auto_score >= 0.88:
            status, conf = "Auto-match", auto_score
            if not sci_code: sci_code,sci_name = auto_code,auto_name
        elif auto_score >= 0.60:
            status, conf = "Sugerido", auto_score
            if not sci_code: sci_code,sci_name = auto_code,auto_name
        elif auto_score > 0:
            status, conf = "Baixa conf.", auto_score
            if not sci_code: sci_code,sci_name = auto_code,auto_name
        else:
            status, conf = "Pendente", 0.0

        rows.append({
            "sigma_code":  nc,
            "sigma_name":  name,
            "sigma_class": cls_sigma,
            "n_movs":      len(movs),
            "total_rs":    round(total,2),
            "sci_code":    sci_code,
            "sci_name":    sci_name,
            "confidence":  round(conf,2),
            "method":      auto_method or source,
            "status":      status,
            "candidates":  candidates,
            "excluir":     excluir,
        })
    return rows


# 
# UTILITÁRIOS
# 

def get_sigma_code(key: str) -> str:
    m = re.match(r"^([\d.]+)", key.strip())
    return m.group(1).rstrip(".") if m else key.strip().rstrip(".")


def date_yyyymmdd(d: str) -> str:
    if not d: return ""
    if "/" in d:
        p = d.split("/")
        if len(p) == 3: return f"{p[2]}{p[1]}{p[0]}"
    digs = re.sub(r"\D","",d)
    return digs if len(digs) == 8 else ""


def date_br(d: str) -> str:
    if len(d) == 8: return f"{d[6:8]}/{d[4:6]}/{d[:4]}"
    return d


def get_nature(sigma_code: str) -> str:
    c = sigma_code.strip()
    if c.startswith("1."): return "D"
    if c.startswith("2."): return "C"
    if c.startswith("3.06") or c.startswith("3.07") or c.startswith("3.08"): return "D"
    if c.startswith("3."): return "C"
    return "D"


def is_excluir(code: str) -> bool:
    if code in EXCLUIR_EXATOS: return True
    return any(code.startswith(p) for p in EXCLUIR_PREFIXOS)


def is_func(text: str) -> bool:
    tu = text.upper()
    return any(f in tu for f in FUNCIONARIOS_REG)


def is_socio(text: str) -> bool:
    return any(s in text.upper() for s in SOCIOS)


def get_sci(sigma_code: str, persistent: dict, extra: dict) -> str:
    nc = get_sigma_code(sigma_code)
    if nc in extra: return extra[nc]
    if nc in persistent: return persistent[nc]
    if nc in DE_PARA_BASE: return DE_PARA_BASE[nc]
    for k,v in DE_PARA_BASE.items():
        if nc.startswith(k+"."): return v
    return ""


def clean_comp(s: str, maxlen: int = 40) -> str:
    if not s: return ""
    return s.replace("\n"," ").replace("\r"," ").replace(",",";").strip()[:maxlen]


def clean_doc(s: str, maxlen: int = 20) -> str:
    if not s: return ""
    return str(s).replace(",",";").replace("\n"," ").strip()[:maxlen]


def assign_dc(code_a, val_a, code_b, val_b):
    nat_a, nat_b = get_nature(code_a), get_nature(code_b)
    if val_a > 0: return (code_a,code_b) if nat_a=="D" else (code_b,code_a)
    return (code_b,code_a) if nat_b=="D" else (code_a,code_b)


# 
# DESDUPLICAÇÃO — 4 PASSOS
# 

def deduplicate(accounts: dict) -> tuple[list,list,dict]:
    all_movs: list[dict] = []
    for key, data in accounts.items():
        nc = get_sigma_code(key)
        if is_excluir(nc): continue
        for m in data.get("movements",[]):
            all_movs.append({
                "code":nc,"name":data.get("name",""),
                "date":m.get("date","") or "","desc":m.get("desc","") or "",
                "doc":m.get("doc","") or "","valor":m.get("valor",0) or 0,
                "abs_valor":abs(m.get("valor",0) or 0),"nature":get_nature(nc),
            })
    paired: set[int] = set()
    lancamentos: list[dict] = []
    alertas: list[dict] = []

    def make_lanc(ia,ib):
        a,b = all_movs[ia],all_movs[ib]
        sd,sc = assign_dc(a["code"],a["valor"],b["code"],b["valor"])
        return {"data":date_yyyymmdd(a["date"]) or date_yyyymmdd(b["date"]) or "",
                "sigma_d":sd,"sigma_c":sc,"valor":a["abs_valor"],
                "desc":a["desc"] or b["desc"],"doc":a["doc"] or b["doc"],
                "tipo":"ADMINISTRATIVO"}

    # P1: mesmo documento
    doc_idx: dict = defaultdict(list)
    for i,m in enumerate(all_movs):
        if m["doc"]: doc_idx[(m["date"],m["doc"],round(m["abs_valor"],2))].append(i)
    for _,idxs in doc_idx.items():
        for i in range(len(idxs)):
            for j in range(i+1,len(idxs)):
                ia,ib = idxs[i],idxs[j]
                if ia in paired or ib in paired: continue
                a,b = all_movs[ia],all_movs[ib]
                if a["code"]==b["code"]: continue
                if a["code"] in ITAU_CODES and b["code"] in ITAU_CODES:
                    paired.add(ia); paired.add(ib); continue
                lancamentos.append(make_lanc(ia,ib)); paired.add(ia); paired.add(ib)

    # P2: data+valor sinais opostos
    val_opp: dict = defaultdict(list)
    for i,m in enumerate(all_movs):
        if i not in paired:
            val_opp[(m["date"],round(m["abs_valor"],2),"pos" if m["valor"]>0 else "neg")].append(i)
    for (dt,val,sign),idxs in val_opp.items():
        opp = [i for i in val_opp.get((dt,val,"neg" if sign=="pos" else "pos"),[]) if i not in paired]
        for ia in idxs:
            if ia in paired: continue
            for ib in opp:
                if ib in paired or ia==ib: continue
                a,b = all_movs[ia],all_movs[ib]
                if a["code"]==b["code"]: continue
                if a["code"] in ITAU_CODES and b["code"] in ITAU_CODES:
                    paired.add(ia); paired.add(ib); break
                lancamentos.append(make_lanc(ia,ib)); paired.add(ia); paired.add(ib); break

    # P3: mesmo DOC, naturezas distintas
    for _,idxs in doc_idx.items():
        unp = [i for i in idxs if i not in paired]
        for i in range(len(unp)):
            for j in range(i+1,len(unp)):
                ia,ib = unp[i],unp[j]
                a,b = all_movs[ia],all_movs[ib]
                if a["nature"]==b["nature"] or a["code"]==b["code"]: continue
                lancamentos.append(make_lanc(ia,ib)); paired.add(ia); paired.add(ib)

    # P4: data+valor mesmo sinal, naturezas distintas
    vs: dict = defaultdict(list)
    for i,m in enumerate(all_movs):
        if i not in paired and not m["doc"]:
            vs[(m["date"],round(m["abs_valor"],2),"pos" if m["valor"]>0 else "neg")].append(i)
    for _,idxs in vs.items():
        for i in range(len(idxs)):
            for j in range(i+1,len(idxs)):
                ia,ib = idxs[i],idxs[j]
                if ia in paired or ib in paired: continue
                a,b = all_movs[ia],all_movs[ib]
                if a["nature"]==b["nature"] or a["code"]==b["code"]: continue
                if a["code"] in ITAU_CODES and b["code"] in ITAU_CODES:
                    paired.add(ia); paired.add(ib); continue
                lancamentos.append(make_lanc(ia,ib)); paired.add(ia); paired.add(ib)

    for i,m in enumerate(all_movs):
        if i not in paired:
            alertas.append({"tipo":"SEM_PAR","code":m["code"],"date":m["date"],
                            "desc":m["desc"],"doc":m["doc"],"valor":m["valor"]})
    return lancamentos, alertas, {
        "total_bruto":len(all_movs),"pares":len(lancamentos),"sem_par":len(alertas)}


# 
# REGRAS DE NEGÓCIO
# 

_NF_RE    = re.compile(r"despesa \(nota\)|nota de servi[cç]o|nf[e\s]\d+|nota fiscal|"
                       r"duplicata|boleto nf|pgto nota|nfs-e|contas a pagar\s*-",re.I)
_CF_RE    = re.compile(r"carta frete|abastec.*frete|frete.*motorista",re.I)
_BONUS_RE = re.compile(r"b[oô]nus|bonifica[cç][aã]o",re.I)
_FOLHA_RE = re.compile(r"sal[aá]rio|folha|fgts|inss|f[eé]rias|13[oº]|d[eé]cimo|"
                       r"rescis[aã]o|pro.labore|holerite|remuner[aá]",re.I)


def apply_rules(lancamentos, accounts, persistent, extra, periodo):
    processados: list[dict] = []
    alertas:     list[dict] = []
    fallback_dt = f"{periodo}01"
    def resolve(sc): return get_sci(sc, persistent, extra)

    for key, data_acc in accounts.items():
        nc   = get_sigma_code(key)
        bsci = BANK_SCI_MAP.get(nc)
        for m in data_acc.get("movements",[]):
            desc=m.get("desc","") or ""; val=m.get("valor",0) or 0
            doc=m.get("doc","") or ""; dt=date_yyyymmdd(m.get("date","") or "") or fallback_dt
            if bsci and re.search(r"sal[aá]rio depositado",desc,re.I):
                processados.append({"data":dt,"sci_d":"1287","sci_c":bsci,
                                    "valor":abs(val),"desc":desc,"doc":doc,"tipo":"FOLHA_A"})
            elif nc=="1.1.2.2.1" and re.search(r"faturamento n[aã]o fiscal",desc,re.I) and val>0 and is_func(desc):
                processados.append({"data":dt,"sci_d":"494050","sci_c":"159",
                                    "valor":abs(val),"desc":desc,"doc":doc,"tipo":"FOLHA_B_CRIA"})
            elif nc=="1.1.2.2.1" and re.search(r"recebimento de faturas desconto de sal[aá]rio",desc,re.I) and val<0 and is_func(desc):
                processados.append({"data":dt,"sci_d":"1287","sci_c":"494050",
                                    "valor":abs(val),"desc":desc,"doc":doc,"tipo":"FOLHA_B_BAIXA"})
            elif nc=="1.1.2.2.2" and re.search(r"vale para motorista",desc,re.I) and val>0 and is_func(desc):
                processados.append({"data":dt,"sci_d":"494051","sci_c":"494018",
                                    "valor":abs(val),"desc":desc,"doc":doc,"tipo":"FOLHA_C_CRIA"})
            elif nc=="1.1.2.2.2" and re.search(r"baixa de notas",desc,re.I) and val<0 and is_func(desc):
                processados.append({"data":dt,"sci_d":"1287","sci_c":"494051",
                                    "valor":abs(val),"desc":desc,"doc":doc,"tipo":"FOLHA_C_BAIXA"})
            elif nc=="1.1.2.3.1" and re.search(r"baixa vale funcionario",desc,re.I) and val<0:
                if is_func(desc):
                    processados.append({"data":dt,"sci_d":"1287","sci_c":"494063",
                                        "valor":abs(val),"desc":desc,"doc":doc,"tipo":"FOLHA_D"})
                elif is_socio(desc):
                    alertas.append({"tipo":"SOCIO_VALE","desc":desc,"valor":abs(val),"date":dt,"doc":doc})

    for l in lancamentos:
        desc=l.get("desc","") or ""; sd=l.get("sigma_d","") or ""
        sc=l.get("sigma_c","") or ""; valor=l.get("valor",0) or 0
        data=l.get("data","") or fallback_dt; doc=l.get("doc","") or ""
        if is_excluir(sd) or is_excluir(sc): continue
        if sd in ITAU_CODES and sc in ITAU_CODES: continue
        if _NF_RE.search(desc):
            bank = sc if get_nature(sc)=="D" else sd
            processados.append({"data":data,"sci_d":"493902","sci_c":resolve(bank),
                                 "valor":valor,"desc":desc,"doc":doc,"tipo":"PASSIVO_FISCAL"})
            continue
        if _FOLHA_RE.search(desc): continue
        if _CF_RE.search(desc):
            dl=desc.lower()
            if "troco" in dl or ("vale" in dl and "motorista" in dl): sd_sci,sc_sci="159","660"
            elif get_nature(sd)=="D" and "1.1.1" in sd: sd_sci,sc_sci=resolve(sd) or "","493909"
            else: sd_sci,sc_sci="493909","159"
            processados.append({"data":data,"sci_d":sd_sci,"sci_c":sc_sci,
                                 "valor":valor,"desc":desc,"doc":doc,"tipo":"CARTA_FRETE"})
            continue
        if _BONUS_RE.search(desc):
            sci_pass = resolve(sc) if get_nature(sc)=="C" else resolve(sd)
            processados.append({"data":data,"sci_d":"494100","sci_c":sci_pass or "",
                                 "valor":valor,"desc":desc,"doc":doc,"tipo":"BONUS"})
            continue
        sci_d = resolve(sd); sci_c = resolve(sc)
        if sci_d=="EXCLUIR" or sci_c=="EXCLUIR": continue
        processados.append({"data":data,"sci_d":sci_d,"sci_c":sci_c,
                             "valor":valor,"desc":desc,"doc":doc,
                             "tipo":l.get("tipo","ADMINISTRATIVO")})
    return processados, alertas


# 
# GERADOR TXT
# 

def gerar_txt(lancamentos: list, periodo: str = "") -> bytes:
    fallback = f"{periodo}01" if periodo else "20260101"
    linhas: list[str] = []
    for i,l in enumerate(sorted(lancamentos,key=lambda x:x.get("data","")),1):
        data  = l.get("data","") or ""
        if len(data) != 8: data = fallback
        sd    = l.get("sci_d","") or ""
        sc    = l.get("sci_c","") or ""
        valor = l.get("valor",0) or 0
        comp  = clean_comp(l.get("desc","") or "")
        ndoc  = clean_doc(l.get("doc","") or "")
        linhas.append(f"{i:06d},{data},{sd},{sc},{valor:.2f},,{comp},{ndoc},,D")
    return ("\r\n".join(linhas)+"\r\n").encode("cp1252",errors="replace")


# 
# GERADOR EXCEL AUDITÁVEL
# 

def gerar_planilha_auditoria(lancamentos: list, sci_plan: dict, per: str) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        rows = [["Seq","Data","Cód Débito","Nome Débito","Cód Crédito",
                 "Nome Crédito","Valor R$","Tipo","Complemento","Documento"]]
        for i,l in enumerate(sorted(lancamentos,key=lambda x:x.get("data","")),1):
            sd=l.get("sci_d",""); sc=l.get("sci_c","")
            rows.append([i,date_br(l.get("data","")),sd,sci_plan.get(sd,""),
                         sc,sci_plan.get(sc,""),
                         f"R$ {l.get('valor',0):,.2f}".replace(",","X").replace(".",",").replace("X","."),
                         l.get("tipo",""),l.get("desc","")[:80],l.get("doc","")])
        buf=io.StringIO(); w=csv.writer(buf,delimiter=";"); w.writerows(rows)
        return buf.getvalue().encode("utf-8-sig")

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = f"Lançamentos {per}"
    thin  = Side(border_style="thin",color="CBD5E1")
    brd   = Border(left=thin,right=thin,top=thin,bottom=thin)
    hfont = Font(bold=True,color="FFFFFF",name="Calibri",size=10)
    hfill = PatternFill(fill_type="solid",fgColor="1A1A2E")
    afill = PatternFill(fill_type="solid",fgColor="EFF6FF")
    ha    = Alignment(horizontal="center",vertical="center",wrap_text=True)
    va    = Alignment(vertical="center")

    headers = ["Seq","Data","Cód. Débito","Nome Débito","Cód. Crédito",
               "Nome Crédito","Valor (R$)","Tipo","Complemento","Documento"]
    widths  = [6,13,12,38,12,38,16,18,45,20]
    for col,(h,w) in enumerate(zip(headers,widths),1):
        cell=ws.cell(row=1,column=col,value=h)
        cell.font=hfont; cell.fill=hfill; cell.alignment=ha; cell.border=brd
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.row_dimensions[1].height=28; ws.freeze_panes="A2"

    for i,l in enumerate(sorted(lancamentos,key=lambda x:x.get("data","")),2):
        sd=l.get("sci_d","") or ""; sc=l.get("sci_c","") or ""
        fill=afill if i%2==0 else None
        row_data=[i-1,date_br(l.get("data","") or ""),sd,sci_plan.get(sd,""),
                  sc,sci_plan.get(sc,""),l.get("valor",0) or 0,
                  l.get("tipo",""),( l.get("desc","") or "")[:80],l.get("doc","") or ""]
        for col,val_cell in enumerate(row_data,1):
            cell=ws.cell(row=i,column=col,value=val_cell)
            cell.alignment=Alignment(horizontal="center" if col in (1,2,3,5,7) else "left",
                                     vertical="center")
            cell.border=brd
            if fill: cell.fill=fill
            if col==7: cell.number_format='#.##0,00'
        ws.row_dimensions[i].height=16

    ws2=wb.create_sheet("Resumo"); tipos=Counter(l.get("tipo","?") for l in lancamentos)
    ws2.append(["Tipo","Qtd","Valor Total R$"])
    for t,n in sorted(tipos.items(),key=lambda x:-x[1]):
        total_t=sum(l.get("valor",0) or 0 for l in lancamentos if l.get("tipo")==t)
        ws2.append([t,n,total_t])
    for col in ("A","B","C"):
        ws2[f"{col}1"].font=Font(bold=True,color="FFFFFF")
        ws2[f"{col}1"].fill=PatternFill(fill_type="solid",fgColor="C9A227")

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# 
# HELPERS DE UI
# 

def metric_card(label: str, value: str, col, gold: bool = False):
    if gold:
        col.markdown(f"""<div class="metric-card-gold">
          <div class="metric-value-gold">{value}</div>
          <div class="metric-label">{label}</div></div>""", unsafe_allow_html=True)
    else:
        col.markdown(f"""<div class="metric-card">
          <div class="metric-value">{value}</div>
          <div class="metric-label">{label}</div></div>""", unsafe_allow_html=True)


# Mapeamento de status → classe CSS do badge
_STATUS_CSS = {
    "Confirmado":  "s-ok",
    "Auto-match":  "s-auto",
    "Sugerido":    "s-sug",
    "Baixa conf.": "s-low",
    "Pendente":    "s-pend",
    "Excluído":    "s-excl",
}
_STATUS_LABEL = {
    "Confirmado":  "Confirmado",
    "Auto-match":  "Auto-match",
    "Sugerido":    "Sugerido",
    "Baixa conf.": "Baixa conf.",
    "Pendente":    "Pendente",
    "Excluído":    "Excluído",
}


def badge_html(status: str) -> str:
    css = _STATUS_CSS.get(status, "s-excl")
    label = _STATUS_LABEL.get(status, status)
    return f'<span class="sbadge {css}">{label}</span>'


def _style_status_df(df):
    """Aplica cores de fundo na coluna Status via pandas Styler."""
    COLORS = {
        "Confirmado":  ("#bbf7d0", "#14532d"),
        "Auto-match":  ("#bfdbfe", "#1e3a8a"),
        "Sugerido":    ("#fef08a", "#713f12"),
        "Baixa conf.": ("#fed7aa", "#7c2d12"),
        "Pendente":    ("#fecaca", "#7f1d1d"),
        "Excluído":    ("#e2e8f0", "#475569"),
    }
    def _sty(val):
        if val in COLORS:
            bg, fg = COLORS[val]
            return (f"background-color:{bg};color:{fg};"
                    f"font-weight:700;border-radius:4px;padding:2px 6px")
        return ""
    cols_exist = [c for c in ["Status"] if c in df.columns]
    if cols_exist:
        return df.style.map(_sty, subset=cols_exist)
    return df.style


# 
# INTERFACE PRINCIPAL
# 

def main():
    #  Hero com logo Domann 
    st.markdown(f"""
    <div class="hero">
      {DOMANN_LOGO_SVG}
      <div style="border-left:1px solid {_DOMANN_GOLD}44;
                  padding-left:1.8rem; margin-left:.6rem;">
        <div style="color:#fff; font-family:'Arial Black',Arial,sans-serif;
                    font-size:1.35rem; font-weight:900; letter-spacing:.6px;
                    text-transform:uppercase;">
          Sistema de Integração Contábil Postos
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

    #  Sidebar 
    with st.sidebar:
        st.markdown("### Configurações")
        empresa_nome = st.text_input("Empresa", value="Auto Posto Murungava Ltda")
        periodo = st.text_input("Período (AAAAMM)", value=datetime.now().strftime("%Y%m"))
        st.divider()
        pers = load_persistent()
        st.caption(f"{len(pers)} mapeamentos confirmados em cache")

        # Exportar mapeamentos (essencial na nuvem) 
        if pers:
            pers_bytes = json.dumps(pers, ensure_ascii=False, indent=2).encode("utf-8")
            st.download_button(
                "Exportar mapeamentos (.json)",
                data=pers_bytes,
                file_name="de_para_confirmado.json",
                mime="application/json",
                use_container_width=True,
                help="Baixe este arquivo e guarde. Na próxima sessão, importe-o para recuperar todos os mapeamentos confirmados.",
            )

        # Importar mapeamentos 
        imp = st.file_uploader("Importar mapeamentos (.json)",
                               type=["json"], label_visibility="collapsed",
                               help="Importe o arquivo de_para_confirmado.json salvo anteriormente.")
        if imp is not None:
            try:
                imported = json.load(imp)
                save_persistent(imported)
                st.success(f"{len(imported)} mapeamentos importados com sucesso.")
                st.rerun()
            except Exception as ex:
                st.error(f"Erro ao importar o arquivo: {ex}")

        if st.button("Limpar sessão", use_container_width=True):
            for k in ["accounts","sci_plan","depara_rows","lancamentos",
                      "stats_dedup","alertas_dedup","alertas_regras"]:
                st.session_state.pop(k,None)
            st.rerun()
        st.divider()
        fe = st.text_area("Funcionários adicionais (maiúsculas):", height=60)
        if fe.strip():
            for fn in fe.strip().upper().splitlines():
                fn=fn.strip()
                if fn: FUNCIONARIOS_REG.add(fn)

    tab1, tab2, tab3, tab4 = st.tabs([
        "Carregar Arquivos",
        "DE/PARA & Mapeamentos",
        "Processamento",
        "Exportar",
    ])

    # 
    # TAB 1 — UPLOAD
    # 
    with tab1:
        st.markdown("### Carregar arquivos do período")
        st.markdown('<div class="gold-box">Carregue a <strong>Movimentação Sigma</strong> '
                    '(obrigatório) e o <strong>Plano de Contas SCI</strong> para ativar o '
                    'motor de matching com lock patrimonial automático.</div>',
                    unsafe_allow_html=True)
        c1,c2 = st.columns(2)
        with c1:
            f_mov = st.file_uploader("Movimentação Sigma *",
                                     type=["xls","xlsx","html","htm","xml","csv"],key="up_mov")
        with c2:
            f_plano = st.file_uploader("Plano de Contas SCI (opcional)",
                                       type=["csv","xlsx","txt"],key="up_plano")
        extra_raw = st.text_area("Mapeamentos manuais (sigma_code=sci_code):",
                                 placeholder="1.1.2.99=494099\n2.1.6.15=494100",height=60)
        extra_de_para: dict[str,str] = {}
        for line in extra_raw.strip().splitlines():
            if "=" in line:
                k,v=line.split("=",1); extra_de_para[k.strip()]=v.strip()

        st.markdown('<hr class="gold-divider">', unsafe_allow_html=True)
        btn_load = st.button("Analisar Arquivos", type="primary", use_container_width=True)

        if btn_load:
            if not f_mov: st.error("Carregue a Movimentação Sigma antes de continuar."); st.stop()
            sci_plan: dict[str,str] = {}
            if f_plano:
                with st.spinner("Lendo Plano de Contas SCI..."):
                    try: sci_plan = parse_sci_plan(f_plano.read(), f_plano.name)
                    except Exception as ex: st.warning(f"Erro SCI: {ex}")
                if sci_plan: st.success(f"Plano de contas SCI carregado: {len(sci_plan)} contas.")
                else: st.warning("Plano SCI não lido — o matching usará apenas o DE/PARA base.")
            with st.spinner(f"Lendo movimentação ({f_mov.name})..."):
                mov_bytes = f_mov.read()
                fmt_det   = _detect_format(mov_bytes, f_mov.name)
                try: accounts = parse_sigma_file(mov_bytes, f_mov.name)
                except Exception as ex:
                    st.error(f"Erro: {ex}")
                    st.code(f"Formato: {fmt_det} | {len(mov_bytes):,} bytes\n{mov_bytes[:300]}")
                    st.stop()
            if not accounts:
                st.error("Não foi possível ler a movimentação. Verifique o formato do arquivo.")
                try: sample=mov_bytes[:600].decode("windows-1252",errors="replace")
                except Exception: sample=str(mov_bytes[:300])
                st.code(sample); st.stop()
            st.session_state.update({
                "accounts":accounts,"sci_plan":sci_plan,"extra_depara":extra_de_para,
                "periodo":periodo,"empresa_nome":empresa_nome,"fmt_detectado":fmt_det,
            })
            for k in ["depara_rows","lancamentos","stats_dedup","alertas_dedup","alertas_regras"]:
                st.session_state.pop(k,None)

        if "accounts" in st.session_state:
            accounts  = st.session_state.accounts
            sci_plan  = st.session_state.get("sci_plan",{})
            n_contas  = len(accounts)
            n_movs    = sum(len(d.get("movements",[])) for d in accounts.values())
            total_val = sum(abs(m.get("valor",0) or 0)
                           for d in accounts.values() for m in d.get("movements",[]))
            st.markdown('<hr class="gold-divider">', unsafe_allow_html=True)
            cc = st.columns(4)
            metric_card("Contas Sigma",  str(n_contas),         cc[0])
            metric_card("Movimentos",    f"{n_movs:,}",         cc[1])
            metric_card("Volume bruto",  f"R$ {total_val:,.0f}",cc[2], gold=True)
            metric_card("Contas SCI",    f"{len(sci_plan)}",    cc[3])
            with st.expander("Prévia das contas Sigma"):
                import pandas as pd
                df = pd.DataFrame([{
                    "Código":get_sigma_code(k),"Classe":detect_class_sigma(get_sigma_code(k)),
                    "Nome da Conta":d.get("name",""),
                    "Movimentos":len(d.get("movements",[])),
                    "Total R$":sum(abs(m.get("valor",0) or 0) for m in d.get("movements",[]))
                } for k,d in sorted(accounts.items())])
                st.dataframe(df, use_container_width=True, hide_index=True)

    # 
    # TAB 2 — DE/PARA & MAPEAMENTOS
    # 
    with tab2:
        if "accounts" not in st.session_state:
            st.info("Carregue os arquivos na aba 1 primeiro."); st.stop()

        accounts   = st.session_state.accounts
        sci_plan   = st.session_state.get("sci_plan",{})
        extra_map  = st.session_state.get("extra_depara",{})
        persistent = load_persistent()

        st.markdown("### Mapeamento DE/PARA — Sigma → SCI")
        st.markdown(
            '<div class="gold-box"><strong>Lock Patrimonial ativo</strong>: '
            'o motor cruza apenas contas da mesma classe (Ativo↔Ativo, Passivo↔Passivo). '
            'Revise as sugestões automáticas abaixo e confirme ou corrija antes de processar.</div>',
            unsafe_allow_html=True)

        if "depara_rows" not in st.session_state:
            with st.spinner("Construindo mapeamentos..."):
                rows = build_depara_with_matches(accounts,sci_plan,persistent,extra_map)
            st.session_state.depara_rows = rows
        else:
            rows = st.session_state.depara_rows

        n_conf = sum(1 for r in rows if r["status"] == "Confirmado")
        n_auto = sum(1 for r in rows if r["status"] == "Auto-match")
        n_sug  = sum(1 for r in rows if r["status"] == "Sugerido")
        n_low  = sum(1 for r in rows if r["status"] == "Baixa conf.")
        n_pend = sum(1 for r in rows if r["status"] == "Pendente")
        n_ex   = sum(1 for r in rows if r["status"] == "Excluído")

        sc1,sc2,sc3,sc4,sc5,sc6 = st.columns(6)
        metric_card("Confirmados", str(n_conf), sc1, gold=True)
        metric_card("Auto-match",  str(n_auto), sc2)
        metric_card("Sugeridos",   str(n_sug),  sc3)
        metric_card("Baixa conf.", str(n_low),  sc4)
        metric_card("Pendentes",   str(n_pend), sc5)
        metric_card("Excluídos",   str(n_ex),   sc6)

        # Filtros
        cf1,cf2 = st.columns([3,2])
        with cf1:
            filtro_status = st.multiselect(
                "Filtrar por status:",
                ["Confirmado","Auto-match","Sugerido","Baixa conf.","Pendente","Excluído"],
                default=["Confirmado","Auto-match","Sugerido","Baixa conf.","Pendente"],
                label_visibility="collapsed")
        with cf2:
            busca = st.text_input("Buscar", value="", placeholder="Buscar por código ou nome...",
                                  label_visibility="collapsed")

        rows_filt = [r for r in rows if not filtro_status or r["status"] in filtro_status]
        if busca and busca != "placeholder":
            bq=busca.lower()
            rows_filt=[r for r in rows_filt
                       if bq in r["sigma_code"].lower() or bq in r["sigma_name"].lower()
                       or bq in (r.get("sci_code") or "").lower()
                       or bq in (r.get("sci_name") or "").lower()]

        # ─── Tabela principal: overview colorido (read-only) ────────────────
        import pandas as pd
        df_overview = pd.DataFrame([{
            "Status":      r["status"],
            "Cód. Sigma":  r["sigma_code"],
            "Nome Sigma":  r["sigma_name"],
            "Classe":      r["sigma_class"],
            "Movs.":       r["n_movs"],
            "Total R$":    r["total_rs"],
            "Cód. SCI":    r["sci_code"],
            "Nome SCI":    r["sci_name"],
            "Conf. %":     int(r["confidence"]*100),
            "Método":      r["method"],
        } for r in rows_filt])

        st.dataframe(
            _style_status_df(df_overview),
            column_config={
                "Status":     st.column_config.TextColumn("Status",      width="small"),
                "Cód. Sigma": st.column_config.TextColumn("Cód. Sigma",  width="small"),
                "Nome Sigma": st.column_config.TextColumn("Nome Sigma",  width="medium"),
                "Classe":     st.column_config.TextColumn("Classe",      width="small"),
                "Movs.":      st.column_config.NumberColumn("Movs.",     width="small"),
                "Total R$":   st.column_config.NumberColumn("Total R$",  format="R$ %.2f"),
                "Cód. SCI":   st.column_config.TextColumn("Cód. SCI",   width="small"),
                "Nome SCI":   st.column_config.TextColumn("Nome SCI",   width="large"),
                "Conf. %":    st.column_config.ProgressColumn("Conf.", min_value=0, max_value=100, width="small"),
                "Método":     st.column_config.TextColumn("Método",     width="small"),
            },
            use_container_width=True, hide_index=True)

        # ─── Edição inline de Código SCI e Nome SCI ─────────────────────────
        st.markdown('<div class="info-box" style="margin-top:.8rem">'
                    'Edite o <strong>Código SCI</strong> e/ou o <strong>Nome SCI</strong> '
                    'diretamente na tabela. Ao salvar, o nome é preenchido automaticamente '
                    'pelo código — ou mantém o nome que você digitou se o código não for '
                    'encontrado no plano. Clique em <strong>Salvar Mapeamentos</strong> '
                    'para confirmar.</div>',
                    unsafe_allow_html=True)

        # Monta df guardando os valores originais para comparação no save
        df_edit_rows = [{
            "Cód. Sigma":      r["sigma_code"],
            "Nome Sigma":      r["sigma_name"],
            "Classe":          r["sigma_class"],
            "Código SCI":      r["sci_code"] or "",
            "Nome SCI":        r["sci_name"] or "",
            "_orig_sci_code":  r["sci_code"] or "",
            "_orig_sci_name":  r["sci_name"] or "",
        } for r in rows_filt]
        df_edit = pd.DataFrame(df_edit_rows)

        edited = st.data_editor(
            df_edit.drop(columns=["_orig_sci_code","_orig_sci_name"]),
            column_config={
                "Cód. Sigma": st.column_config.TextColumn("Cód. Sigma",  width="small",  disabled=True),
                "Nome Sigma": st.column_config.TextColumn("Nome Sigma",  width="medium", disabled=True),
                "Classe":     st.column_config.TextColumn("Classe",      width="small",  disabled=True),
                "Código SCI": st.column_config.TextColumn("Código SCI",  width="small",
                                  help="Digite ou corrija o código SCI. O nome será auto-preenchido ao salvar."),
                "Nome SCI":   st.column_config.TextColumn("Nome SCI",    width="large",
                                  help="Preenchido automaticamente pelo código. Você pode sobrescrever manualmente."),
            },
            use_container_width=True, hide_index=True, num_rows="fixed",
            key="depara_editor")

        cs1,cs2 = st.columns([1,3])
        with cs1:
            btn_save = st.button("Salvar Mapeamentos", type="primary", use_container_width=True)
        with cs2:
            if btn_save:
                st.markdown('<div class="ok-box" style="margin:0">Salvo em '
                            '<code>de_para_confirmado.json</code> — '
                            'aplicado automaticamente nos próximos períodos.</div>',
                            unsafe_allow_html=True)

        if btn_save:
            to_save: dict[str,str] = {}
            edit_map: dict[str,tuple[str,str]] = {}  # sigma_code → (sci_code, sci_name)

            for idx, row in edited.iterrows():
                sg        = str(row["Cód. Sigma"]).strip()
                sc_code   = str(row["Código SCI"]).strip()
                sc_name_u = str(row["Nome SCI"]).strip()   # nome digitado pelo usuário
                orig_code = str(df_edit_rows[idx]["_orig_sci_code"]).strip()
                orig_name = str(df_edit_rows[idx]["_orig_sci_name"]).strip()

                if not sg or sc_code in ("nan",""):
                    continue

                # Auto-preenchimento: se código mudou → busca nome no plano
                if sc_code != orig_code:
                    resolved_name = sci_plan.get(sc_code, sc_name_u)
                else:
                    # Código igual: respeita nome digitado pelo usuário, ou mantém original
                    resolved_name = sc_name_u if sc_name_u not in ("nan","",orig_name) else orig_name

                edit_map[sg] = (sc_code, resolved_name)
                if sc_code not in ("EXCLUIR",):
                    to_save[sg] = sc_code

            save_persistent(to_save)

            new_rows=[]
            for r in st.session_state.depara_rows:
                nr=dict(r)
                if r["sigma_code"] in edit_map:
                    ec, en = edit_map[r["sigma_code"]]
                    if ec and ec not in ("nan",""):
                        nr["sci_code"]=ec; nr["sci_name"]=en
                        nr["status"]="Confirmado"; nr["confidence"]=1.0
                new_rows.append(nr)
            st.session_state.depara_rows=new_rows
            st.success(f"{len(to_save)} mapeamentos salvos com sucesso.")

        #  SEÇÃO: ALTERAR SUGESTÃO (para corrigir sugestões erradas) 
        st.markdown('<hr class="gold-divider">', unsafe_allow_html=True)
        altera_rows = [r for r in rows
                       if r["status"] in ("Auto-match","Sugerido","Baixa conf.")
                       and (r.get("candidates") or sci_plan)]

        if altera_rows:
            with st.expander(
                f"Corrigir Sugestão Automática — {len(altera_rows)} contas disponíveis",
                expanded=(n_auto+n_sug > 0 and n_pend == 0)):
                st.markdown(
                    '<div class="warn-box">O sistema sugeriu automaticamente as contas '
                    'abaixo. Verifique se a conta SCI está correta. '
                    'Se não estiver, selecione a conta correta e clique em '
                    '<strong>Confirmar</strong>.</div>',
                    unsafe_allow_html=True)

                sci_by_class: dict[str,list] = defaultdict(list)
                if sci_plan:
                    for code,name in sorted(sci_plan.items(),key=lambda x:x[1]):
                        cls=detect_class_sci(name)
                        sci_by_class[cls].append((code,name))
                    sci_by_class["TODOS"] = sorted(sci_plan.items(),key=lambda x:x[1])

                confirmed_changes: dict[str,str] = {}
                for r in altera_rows:
                    cls_s = r["sigma_class"]
                    curr_code = r["sci_code"]
                    curr_name = r["sci_name"]
                    cands = r.get("candidates",[])

                    st.markdown(
                        f'<div class="alt-card">'
                        f'<strong>{r["sigma_code"]}</strong> — {r["sigma_name"]} '
                        f'<small>({cls_s})</small><br>'
                        f'<small>Sugestão atual: <code>{curr_code}</code> '
                        f'{curr_name} &nbsp;|&nbsp; '
                        f'Conf.: {int(r["confidence"]*100)}% ({r["method"]})</small></div>',
                        unsafe_allow_html=True)

                    col_a, col_b = st.columns([3,1])
                    with col_a:
                        if sci_plan:
                            # Opções filtradas pela classe
                            opts_cls = sci_by_class.get(cls_s,[]) + sci_by_class.get("OUTROS",[])
                            if not opts_cls: opts_cls = sci_by_class.get("TODOS",[])
                            options_display = [f"{c} — {n}" for c,n in opts_cls]
                            # Pré-selecionar atual
                            pre_opt = f"{curr_code} — {curr_name}" if curr_name else ""
                            pre_idx = options_display.index(pre_opt)+1 if pre_opt in options_display else 0
                            sel = st.selectbox(
                                f"Nova conta SCI ({cls_s}):",
                                ["— manter atual —"] + options_display,
                                index=pre_idx,
                                key=f"alt_{r['sigma_code']}",
                                help=f"Contas filtradas para classe {cls_s}. Digite para pesquisar.")
                        else:
                            sel = st.text_input(
                                "Código SCI correto:",
                                value=curr_code,
                                key=f"alt_{r['sigma_code']}",
                                placeholder="Ex: 494018")
                            sel = sel if sel and sel != curr_code else "— manter atual —"

                        if cands and len(cands) > 1:
                            st.caption("**Top 5 candidatos por classe:**  " +
                                       "  |  ".join(f"`{c}` {n[:25]} ({int(s*100)}%)"
                                                    for c,n,s,_ in cands[:5]))
                    with col_b:
                        if st.button("Confirmar", key=f"btn_alt_{r['sigma_code']}",
                                     use_container_width=True):
                            if sel not in ("— manter atual —",""):
                                code_sel = sel.split(" — ")[0].strip() if " — " in sel else sel.strip()
                                if code_sel:
                                    save_persistent({r["sigma_code"]: code_sel})
                                    confirmed_changes[r["sigma_code"]] = code_sel
                                    st.success(f"Mapeamento salvo: {r['sigma_code']} → {code_sel}")

                if confirmed_changes:
                    new_rows=[]
                    for r in st.session_state.depara_rows:
                        nr=dict(r)
                        if r["sigma_code"] in confirmed_changes:
                            nr["sci_code"]=confirmed_changes[r["sigma_code"]]
                            nr["sci_name"]=sci_plan.get(nr["sci_code"],"")
                            nr["status"]="Confirmado"; nr["confidence"]=1.0
                        new_rows.append(nr)
                    st.session_state.depara_rows=new_rows

        # ─── SEÇÃO: RESOLVER PENDÊNCIAS ──────────────────────────────────────
        pendentes = [r for r in rows if r["status"] == "Pendente"]
        if pendentes:
            st.markdown('<hr class="gold-divider">', unsafe_allow_html=True)
            with st.expander(f"Resolver Pendências — {len(pendentes)} contas sem mapeamento"):
                st.markdown(
                    '<div class="warn-box">Contas Sigma sem correspondência no DE/PARA base. '
                    'Selecione a conta SCI equivalente e confirme para cada uma.</div>',
                    unsafe_allow_html=True)
                if not sci_plan:
                    st.info("Carregue o Plano de Contas SCI na aba 1 para ativar a busca.")
                else:
                    sci_by_class2: dict[str,list] = defaultdict(list)
                    for code,name in sorted(sci_plan.items(),key=lambda x:x[1]):
                        sci_by_class2[detect_class_sci(name)].append((code,name))
                    sci_by_class2["TODOS"]=sorted(sci_plan.items(),key=lambda x:x[1])

                    confirmed_pend: dict[str,str] = {}
                    for r in pendentes:
                        cls_s = r["sigma_class"]
                        st.markdown(
                            f'<div class="pend-card">'
                            f'<strong>{r["sigma_code"]}</strong> — {r["sigma_name"]} '
                            f'<small>({cls_s} | {r["n_movs"]} movs | '
                            f'R$ {r["total_rs"]:,.2f})</small></div>',
                            unsafe_allow_html=True)
                        opts_cls=sci_by_class2.get(cls_s,[])+sci_by_class2.get("OUTROS",[])
                        if not opts_cls: opts_cls=sci_by_class2.get("TODOS",[])
                        options_display=[f"{c} — {n}" for c,n in opts_cls]
                        cands=r.get("candidates",[])
                        pre=""
                        if cands: pre=f"{cands[0][0]} — {cands[0][1]}"
                        pre_idx=options_display.index(pre)+1 if pre in options_display else 0
                        pca,pcb=st.columns([4,1])
                        with pca:
                            sel_p=st.selectbox(
                                f"Conta SCI para {r['sigma_code']}:",
                                ["— Selecione —"]+options_display,
                                index=pre_idx,
                                key=f"pend_{r['sigma_code']}")
                        with pcb:
                            if st.button("Confirmar", key=f"btn_pend_{r['sigma_code']}",
                                         use_container_width=True):
                                if sel_p != "— Selecione —":
                                    code_sel=sel_p.split(" — ")[0].strip()
                                    save_persistent({r["sigma_code"]:code_sel})
                                    confirmed_pend[r["sigma_code"]]=code_sel
                                    st.success(f"{r['sigma_code']} → {code_sel}")
                    if confirmed_pend:
                        new_rows=[]
                        for r in st.session_state.depara_rows:
                            nr=dict(r)
                            if r["sigma_code"] in confirmed_pend:
                                nr["sci_code"]=confirmed_pend[r["sigma_code"]]
                                nr["sci_name"]=sci_plan.get(nr["sci_code"],"")
                                nr["status"]="Confirmado"; nr["confidence"]=1.0
                            new_rows.append(nr)
                        st.session_state.depara_rows=new_rows

    # 
    # TAB 3 — PROCESSAMENTO
    # 
    with tab3:
        if "accounts" not in st.session_state:
            st.info("Carregue os arquivos na aba 1 primeiro."); st.stop()
        accounts   = st.session_state.accounts
        per        = st.session_state.get("periodo", periodo)
        persistent = load_persistent()
        extra_map  = st.session_state.get("extra_depara",{})
        depara_editado: dict[str,str] = {}
        if "depara_rows" in st.session_state:
            for r in st.session_state.depara_rows:
                if r.get("sci_code") and r["sci_code"] not in ("EXCLUIR",""):
                    depara_editado[r["sigma_code"]]=r["sci_code"]
        merged_extra={**extra_map,**depara_editado}

        st.markdown("### Processamento Contábil")
        btn_proc=st.button("Executar Processamento", type="primary", use_container_width=True)

        if btn_proc or "lancamentos" in st.session_state:
            if btn_proc:
                with st.spinner("Deduplicando (4 passos)..."):
                    lancs_dedup,alertas_dedup,stats=deduplicate(accounts)
                with st.spinner("Aplicando regras contábeis..."):
                    lancamentos,alertas_regras=apply_rules(lancs_dedup,accounts,persistent,merged_extra,per)
                st.session_state.update({
                    "lancamentos":lancamentos,"alertas_dedup":alertas_dedup,
                    "alertas_regras":alertas_regras,"stats_dedup":stats})

            lancamentos   = st.session_state.lancamentos
            alertas_dedup = st.session_state.alertas_dedup
            alertas_regras= st.session_state.get("alertas_regras",[])
            stats         = st.session_state.stats_dedup
            completos=[l for l in lancamentos if l.get("sci_d") and l.get("sci_c")]
            pend_l   =[l for l in lancamentos if not l.get("sci_d") or not l.get("sci_c")]
            tipos    =Counter(l.get("tipo","?") for l in lancamentos)

            pm=st.columns(5)
            metric_card("Movs. brutos", f"{stats['total_bruto']:,}",pm[0])
            metric_card("Pares gerados",f"{stats['pares']:,}",     pm[1])
            metric_card("Sem par",      f"{stats['sem_par']:,}",   pm[2])
            metric_card("TXT completos",f"{len(completos):,}",     pm[3],gold=True)
            metric_card("Pendências",   f"{len(pend_l)}",          pm[4])

            st.markdown('<hr class="gold-divider">', unsafe_allow_html=True)
            import pandas as pd
            tc1,tc2=st.columns(2)
            with tc1:
                st.markdown("#### Distribuição por tipo")
                st.dataframe(pd.DataFrame(
                    [{"Tipo":t,"Qtd":n,"%":f"{n/max(len(lancamentos),1)*100:.1f}%"}
                     for t,n in sorted(tipos.items(),key=lambda x:-x[1])]),
                    use_container_width=True,hide_index=True)
            with tc2:
                st.markdown("#### Pendências D/C")
                if pend_l:
                    st.dataframe(pd.DataFrame([{
                        "Data":date_br(l.get("data","")),"Falta":"Débito" if not l.get("sci_d") else "Crédito",
                        "Valor":l.get("valor",0),"Desc":(l.get("desc") or "")[:40],
                    } for l in pend_l[:50]]),use_container_width=True,hide_index=True)
                else:
                    st.markdown('<div class="ok-box">Todos os lançamentos têm débito e crédito preenchidos.</div>',
                                unsafe_allow_html=True)

            if alertas_dedup:
                with st.expander(f"Movimentos sem par — {len(alertas_dedup)} itens para revisão"):
                    st.markdown('<div class="warn-box">Movimentos que não encontraram contrapartida no ''período. Verifique se há lançamentos incompletos no Sigma.</div>',
                                unsafe_allow_html=True)
                    st.dataframe(pd.DataFrame([{
                        "Data":a["date"],"Conta":a["code"],
                        "Valor":a["valor"],"Descricao":(a["desc"] or "")[:50]}
                        for a in alertas_dedup[:100]]),use_container_width=True,hide_index=True)
            if alertas_regras:
                with st.expander(f"Alertas de socios — {len(alertas_regras)} itens"):
                    for a in alertas_regras:
                        st.markdown(f"- **{a.get('tipo')}** | {a.get('date')} | "
                                    f"R$ {a.get('valor',0):,.2f} | {a.get('desc','')[:60]}")
            with st.expander("Preview dos lancamentos (primeiros 100)"):
                st.dataframe(pd.DataFrame([{
                    "Data":date_br(l["data"]),"Tipo":l.get("tipo",""),
                    "D":l.get("sci_d",""),"C":l.get("sci_c",""),
                    "Valor":l["valor"],"Desc":(l.get("desc") or "")[:38],
                } for l in lancamentos[:100]]),use_container_width=True,hide_index=True)

    # TAB 4 - EXPORTAR
    with tab4:
        if "lancamentos" not in st.session_state:
            st.info("Execute o processamento na aba 3 primeiro."); st.stop()
        lancamentos = st.session_state.lancamentos
        per         = st.session_state.get("periodo", periodo)

        st.markdown(f'<h2 class="section-h">Exportar TXT para o SCI</h2>', unsafe_allow_html=True)

        col_exp1, col_exp2 = st.columns([2,1])
        with col_exp1:
            st.markdown(f"""
<div class="info-card">
<strong>Periodo:</strong> {per} &nbsp;|&nbsp;
<strong>Lancamentos:</strong> {len(lancamentos)} &nbsp;|&nbsp;
<strong>Formato:</strong> Windows-1252 (ANSI)
</div>""", unsafe_allow_html=True)

        total_d = sum(l["valor"] for l in lancamentos if l.get("sci_d") and l.get("sci_c"))
        total_c = sum(l["valor"] for l in lancamentos if l.get("sci_d") and l.get("sci_c"))
        pend    = sum(1 for l in lancamentos if not l.get("sci_d") or not l.get("sci_c"))

        if pend:
            st.markdown(
                f'<div class="warn-box">Atencao: <strong>{pend}</strong> lancamentos com debito ou credito em branco (pendentes). '
                f'Revise antes de importar no SCI.</div>',
                unsafe_allow_html=True)

        if total_d == total_c:
            st.markdown(
                f'<div class="ok-box">Balanceado: Total debitos = Total creditos = R$ {total_d:,.2f}</div>',
                unsafe_allow_html=True)

        # Gerar TXT
        txt_lines = []
        seq = 1
        for l in lancamentos:
            d   = l.get("sci_d","")
            c   = l.get("sci_c","")
            val = f"{l['valor']:.2f}"
            dt  = l["data"].replace("-","")
            desc = clean_comp(l.get("desc",""), 40)
            seq_str = str(seq).zfill(6)
            txt_lines.append(f"{seq_str},{dt},{d},{c},{val},,,{desc},,,D")
            seq += 1

        txt_content = "\n".join(txt_lines)

        with col_exp2:
            st.download_button(
                label="Baixar TXT para o SCI",
                data=txt_content.encode("windows-1252", errors="replace"),
                file_name=f"importacao_sci_{per.replace('/','_')}.txt",
                mime="text/plain",
                use_container_width=True,
            )

        with st.expander("Visualizar TXT gerado"):
            st.code(txt_content[:3000] + ("\n..." if len(txt_content) > 3000 else ""), language="text")

        # Exportar DE/PARA confirmado
        st.markdown("---")
        st.markdown('<h3 class="section-h" style="font-size:1rem">Salvar DE/PARA confirmado</h3>',
                    unsafe_allow_html=True)
        col_j1, col_j2 = st.columns([2,1])
        with col_j1:
            st.caption("Baixe e salve este arquivo junto com o app para preservar seus mapeamentos.")
        with col_j2:
            pers_export = load_persistent()
            st.download_button(
                label="Baixar de_para_confirmado.json",
                data=json.dumps(pers_export, ensure_ascii=False, indent=2).encode("utf-8"),
                file_name="de_para_confirmado.json",
                mime="application/json",
                use_container_width=True,
            )

